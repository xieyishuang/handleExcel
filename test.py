import openpyxl
import os

def run_summary():
    path = os.getcwd()
    for filename in os.listdir(path):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            filepath =path + os.sep + filename

    wb = openpyxl.open(filepath,read_only=True)
    sheets = wb.sheetnames

    lst = []
    months = []
    dic = {}

    for item in sheets:
        print(item)
        if '月' in item:
            months.append(item)
            sheet = wb[item]
            rows = sheet.max_row
            dic[item] = {}
            for i in range(1,rows+1):
                if sheet.cell(i, 1).font.bold == False:
                    sign = sheet.cell(row = i, column = 1).value.replace(' ','')
                    lst.append(sign)
                    dic[item][sign] = int(sheet.cell(row = i, column = 2).value)
    lst = set(lst)
    lst = list(lst)

    nb = openpyxl.Workbook()
    ns = nb.active
    ns.cell(1,1).value = '签名'
    for row in range(1, len(lst)+1):
        ns.cell(row+1,1).value = lst[row-1]

    for col in range(0,len(months)):
        ns.cell(1,col+2).value = months[col]

    rows = ns.max_row
    cols = ns.max_column

    for col in range(cols):
        if ns.cell(1,col+1).value in dic.keys():
            for row in range(rows):
                if ns.cell(row+1,1).value in dic[ns.cell(1,col+1).value].keys():
                    ns.cell(row+1,col+1).value = dic[ns.cell(1,col+1).value][ns.cell(row+1,1).value]


    save_path = path +os.sep + 'result.xlsx'
    nb.save(save_path)




if __name__ == '__main__':
    run_summary()

