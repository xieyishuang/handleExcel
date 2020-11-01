import openpyxl
#打开Excel
wb = openpyxl.load_workbook('C:\\Users\yishuang.xie\Desktop\云峰3-8月分摊.xlsx')
# 获取sheet
sheet = wb['3月份']
# print (sheet['A6'].value)

listColumn = []
listRow = []
rows = []

for row in sheet.iter_rows():
            rows.append(row)
print(rows)

# for i in range(sheet.max_row):
    # for j in range(sheet.max_column):
        # v = rows[i][j].value
        # print(type(v))
        # if "公司" not in rows[i][j].value:
        # listColumn.append(rows[i][j].value)

for i in range(sheet.max_row):
    v = rows[i][0].value
    # print(type(v))
    if "公司" not in rows[i][0].value:
        listColumn.append(v)
# for i in range(sheet.max_row):
#     listColumn.append(rows[i][1].value)

aaa=[listColumn[i::sheet.max_column] for i in range(sheet.max_column)]

print(aaa)
