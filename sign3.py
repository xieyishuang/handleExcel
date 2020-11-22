import openpyxl
from openpyxl.styles import fonts
#打开Excel
# wb = openpyxl.load_workbook('C:\\Users\yishuang.xie\Desktop\云峰3-8月分摊.xlsx')
wb = openpyxl.load_workbook('F:\\python处理excel例子\云峰3-8月分摊.xlsx')
# 根据sheet名获取sheet
# sheet = wb['3月份']
# allSheet = wb.get_sheet_names
allSheet = wb.sheetnames

# filepath = 'new_excel.xlsx'
# wb1 = openpyxl.Workbook()
# 默认表sheet1
wb1 = openpyxl.load_workbook('F:\\PycharmProject\handleExcel\\new_excel.xlsx')
ws1 = wb1.active
# 更改表名
# ws1.title = 'new_sheet_name'
# 创建sheet2表
# ws2 = wb.create_sheet('sheet2')

for j in range(len(allSheet)-5):
# for j in range(0,1):
    print(allSheet[j])
    sheet = wb[allSheet[j]]
    tempList = []
    for row in sheet.rows:
        for cell in row:
            # cell.coordinate
            if cell.value is not None and cell.font.b == False:
                tempList.append(cell.value)
                # print(cell.value)

# for i in tempList:
#     print(i)

    # for cell in ws1.columns[0]:
    #     if tempList[::3][i] == cell.value:

    maxRow = int(len(tempList)/3)

    count = 0
    for i in range(0,maxRow):
        # for row in ws1.rows:
        #     for cell in row:
        # for column in ws1.columns[0]:
        isHave = False

        for cell in ws1["A"]:

            # if cell.value is not None:
            #     count += 1
            if tempList[::3][i] == cell.value:
                isHave = True
                break
                # pass
                # rowColumn = cell.coordinate
                # ws1.cell(row=(cell.coordinate)[1::], column=j+2, value=tempList[1::3][i])

            else:
                # ws1.cell(row=count + 1, column=j+1, value=tempList[::3][i])
                # ws1.cell(row=count + 1, column=j+2, value=tempList[1::3][i])
                isHave = False
                # if cell.value is not None:
                #     count += 1
                # count += 1
        # ws1.cell(row=i,column=1,value=tempList[::3][i - 1]).value
        # ws1.cell(row=i,column=2,value=tempList[1::3][i - 1]).value

        if j == 0 and i == 0:
            startRow1 = 1
        else:
            startRow1 = ws1.max_row+1
        if not isHave:
            ws1.cell(row=startRow1, column=1, value=tempList[::3][i])
            ws1.cell(row=startRow1, column=j + 2, value=tempList[1::3][i])
        else:
            # ws1.cell(row=int((cell.coordinate)[1::]), column=1, value=tempList[::3][i])
            ws1.cell(row=int((cell.coordinate)[1::]), column=j + 2, value=tempList[1::3][i])
wb1.save("F:\\PycharmProject\handleExcel\\new_excel.xlsx")
# wb1.save(filepath)



# print(sheet['A6'].font.b)
# listColumn = []
# listRow = []
# rows = []
# for row in sheet.iter_rows():
#     if row.font.b == false:
#         rows.append(row)
# print(rows)
# for i in range(sheet.max_row):
#     v = rows[i][0].value
#     print()
    # if rows[i][0].value.font.b == false:
    # if "公司" not in rows[i][0].value:
#     listColumn.append(v)
# for i in listColumn:
#     print(i)
# print(len(listColumn))
# aaa=[listColumn[i::sheet.max_column] for i in range(sheet.max_column)]
# aaa=[listColumn[i::sheet.max_column] for i in range(sheet.max_column)]

# print(aaa)
