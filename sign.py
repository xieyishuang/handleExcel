import openpyxl
import os
from openpyxl.styles import fonts
#打开Excel
# wb = openpyxl.load_workbook('C:\\Users\yishuang.xie\Desktop\云峰3-8月分摊.xlsx')
inputPath = os.getcwdb()
print("%s"%inputPath)
wb = openpyxl.load_workbook(str(inputPath,encoding="utf-8")+'\云峰3-8月分摊.xlsx',read_only=True)
# wb = openpyxl.load_workbook('F:\\python处理excel例子\云峰3-8月分摊.xlsx')
# 根据sheet名获取sheet
# sheet = wb['3月份']
# allSheet = wb.get_sheet_names
allSheet = wb.sheetnames

filepath = 'new_excel.xlsx'
wb1 = openpyxl.Workbook()
# 默认表sheet1
# wb1 = openpyxl.load_workbook('F:\\PycharmProject\handleExcel\\new_excel.xlsx')
ws1 = wb1.active
# 更改表名
# ws1.title = 'new_sheet_name'
# 创建sheet2表
# ws2 = wb.create_sheet('sheet2')

for j in range(len(allSheet)-5):
# for j in range(0,1):
    print(allSheet[j])
    sheet = wb[allSheet[j]]
    tempList = []
    for row in sheet.rows:
        for cell in row:
            # cell.coordinate
            if cell.value is not None and cell.font.b == False:
                tempList.append(cell.value)
                # print(cell.value)
    maxRow = int(len(tempList)/3)
    count = 0
    for i in range(0,maxRow):
        isHave = False

        for cell in ws1["A"]:
            if tempList[::3][i] == cell.value:
                isHave = True
                break

            else:
                isHave = False
        if j == 0 and i == 0:
            startRow1 = 1
        else:
            startRow1 = ws1.max_row+1
        if not isHave:
            ws1.cell(row=startRow1, column=1, value=tempList[::3][i])
            ws1.cell(row=startRow1, column=j + 2, value=tempList[1::3][i])
        else:
            # ws1.cell(row=int((cell.coordinate)[1::]), column=1, value=tempList[::3][i])
            ws1.cell(row=int((cell.coordinate)[1::]), column=j + 2, value=tempList[1::3][i])
# wb1.save("F:\\PycharmProject\handleExcel\\new_excel.xlsx")
wb1.save(filepath)

